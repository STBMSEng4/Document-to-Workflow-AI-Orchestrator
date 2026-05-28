"""Tests for structured table export helpers."""

from __future__ import annotations

from io import BytesIO
import unittest

from openpyxl import load_workbook

from app.exporters import flatten_records, rows_to_csv_bytes, tables_to_xlsx_bytes
from app.schemas import EquipmentRecord, PointRecord


class TabularExporterTests(unittest.TestCase):
    def test_flatten_records_expands_nested_attributes(self) -> None:
        equipment = EquipmentRecord(
            equipment_tag="RTU-1",
            equipment_type="RTU",
            equipment_name="Roof Top Unit 1",
            source_reference="Schedule A / Table 1",
            confidence=0.95,
            attributes={"electrical": {"voltage": "480V/3PH"}},
        )

        rows = flatten_records([equipment])

        self.assertEqual(rows[0]["equipment_tag"], "RTU-1")
        self.assertEqual(rows[0]["attributes.electrical.voltage"], "480V/3PH")

    def test_rows_to_csv_bytes_contains_headers(self) -> None:
        point = PointRecord(
            equipment_tag="AHU-1",
            equipment_type="AHU",
            point_name="Supply Air Temperature",
            point_code="SAT",
            point_type="AI",
            signal_type="analog",
            point_role="sensor",
            engineering_unit="degF",
            source_reference="Sheet P-1",
            confidence=0.95,
            source_type="source_extracted",
        )

        rows = flatten_records([point])
        csv_text = rows_to_csv_bytes(rows).decode("utf-8")

        self.assertIn("equipment_tag", csv_text)
        self.assertIn("Supply Air Temperature", csv_text)

    def test_tables_to_xlsx_bytes_writes_multiple_sheets(self) -> None:
        equipment_rows = [{"equipment_tag": "RTU-1", "equipment_type": "RTU"}]
        point_rows = [{"equipment_tag": "RTU-1", "point_name": "Supply Fan Status"}]

        workbook_bytes = tables_to_xlsx_bytes(
            {"Equipment": equipment_rows, "Points": point_rows}
        )

        workbook = load_workbook(BytesIO(workbook_bytes))
        self.assertIn("Equipment", workbook.sheetnames)
        self.assertIn("Points", workbook.sheetnames)
        self.assertEqual(workbook["Equipment"]["A2"].value, "RTU-1")


if __name__ == "__main__":
    unittest.main()
