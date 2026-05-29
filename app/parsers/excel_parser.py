"""Tabular parser for SpecFlow AI.

Converts .xlsx / .xls / .csv files to Markdown tables.
Designed for points lists, I/O schedules, and equipment schedules.
"""

from __future__ import annotations

import csv
import hashlib
from datetime import datetime, timezone
from pathlib import Path

OUTPUTS_RAW = Path(__file__).parents[2] / "outputs" / "markdown" / "raw"


def _sheet_to_markdown(ws) -> str:
    """Convert an openpyxl worksheet to a Markdown table."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    # Find the last non-empty column
    max_col = 0
    for row in rows:
        for i, cell in enumerate(row):
            if cell is not None and str(cell).strip():
                max_col = max(max_col, i + 1)

    if max_col == 0:
        return ""

    lines: list[str] = []
    header_done = False
    for row in rows:
        cells = [str(cell).strip() if cell is not None else "" for cell in row[:max_col]]
        # Skip completely empty rows
        if not any(cells):
            continue
        lines.append("| " + " | ".join(cells) + " |")
        if not header_done:
            lines.append("|" + "|".join(["---"] * max_col) + "|")
            header_done = True

    return "\n".join(lines)


def _workbook_to_markdown(path: Path) -> str:
    """Convert all sheets of an xlsx file to Markdown."""
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md = _sheet_to_markdown(ws)
        if md:
            sections.append(f"## Sheet: {sheet_name}\n\n{md}")

    wb.close()
    return "\n\n".join(sections)


def _csv_to_markdown(path: Path) -> str:
    """Convert a CSV file to a Markdown table."""
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue

    if not text.strip():
        return ""

    reader = csv.reader(text.splitlines())
    rows = [[cell.strip() for cell in row] for row in reader]
    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        return ""

    max_col = max(len(row) for row in rows)
    normalized_rows = [row + [""] * (max_col - len(row)) for row in rows]

    lines: list[str] = [f"## Sheet: {path.stem}", ""]
    for index, row in enumerate(normalized_rows):
        escaped = [cell.replace("|", "\\|") for cell in row]
        lines.append("| " + " | ".join(escaped) + " |")
        if index == 0:
            lines.append("|" + "|".join(["---"] * max_col) + "|")
    return "\n".join(lines)


def _classify_excel(markdown: str) -> str:
    if len(markdown.strip()) < 20:
        return "empty_or_failed_parse"
    return "text_pdf"


def parse_excel_to_markdown(file_path: str) -> dict:
    """Parse a .xlsx file and return SpecFlow-compatible result dict."""
    path = Path(file_path)
    errors: list[str] = []
    raw_markdown = ""
    classification = "empty_or_failed_parse"

    if not path.exists():
        return {
            "source_type": "xlsx",
            "source_file": str(file_path),
            "ingestion_engine": "openpyxl",
            "pdf_classification": "empty_or_failed_parse",
            "raw_markdown": "",
            "metadata": {},
            "ocr_required": False,
            "status": "failed",
            "errors": [f"File not found: {file_path}"],
        }

    try:
        raw_markdown = _workbook_to_markdown(path)
        classification = _classify_excel(raw_markdown)
    except ImportError:
        errors.append("openpyxl not installed. Run: pip install openpyxl")
        classification = "empty_or_failed_parse"
    except Exception as exc:
        errors.append(f"Excel extraction error: {exc}")
        classification = "empty_or_failed_parse"

    file_hash = hashlib.md5(path.read_bytes()).hexdigest() if path.exists() else ""
    metadata = {
        "source_file": path.name,
        "file_size_bytes": path.stat().st_size if path.exists() else 0,
        "file_hash_md5": file_hash,
        "ingested_at": datetime.now(timezone.utc).isoformat(),
        "pdf_classification": classification,
        "char_count": len(raw_markdown),
        "word_count": len(raw_markdown.split()),
    }

    if raw_markdown:
        OUTPUTS_RAW.mkdir(parents=True, exist_ok=True)
        out_path = OUTPUTS_RAW / f"{path.stem}_raw.md"
        out_path.write_text(raw_markdown, encoding="utf-8")

    status = "failed" if not raw_markdown.strip() else "success"

    return {
        "source_type": "xlsx",
        "source_file": path.name,
        "ingestion_engine": "openpyxl",
        "pdf_classification": classification,
        "raw_markdown": raw_markdown,
        "metadata": metadata,
        "ocr_required": False,
        "status": status,
        "errors": errors,
    }


def parse_csv_to_markdown(file_path: str) -> dict:
    """Parse a .csv file and return SpecFlow-compatible result dict."""
    path = Path(file_path)
    errors: list[str] = []
    raw_markdown = ""
    classification = "empty_or_failed_parse"

    if not path.exists():
        return {
            "source_type": "csv",
            "source_file": str(file_path),
            "ingestion_engine": "python-csv",
            "pdf_classification": "empty_or_failed_parse",
            "raw_markdown": "",
            "metadata": {},
            "ocr_required": False,
            "status": "failed",
            "errors": [f"File not found: {file_path}"],
        }

    try:
        raw_markdown = _csv_to_markdown(path)
        classification = _classify_excel(raw_markdown)
    except Exception as exc:
        errors.append(f"CSV extraction error: {exc}")
        classification = "empty_or_failed_parse"

    file_hash = hashlib.md5(path.read_bytes()).hexdigest() if path.exists() else ""
    metadata = {
        "source_file": path.name,
        "file_size_bytes": path.stat().st_size if path.exists() else 0,
        "file_hash_md5": file_hash,
        "ingested_at": datetime.now(timezone.utc).isoformat(),
        "pdf_classification": classification,
        "char_count": len(raw_markdown),
        "word_count": len(raw_markdown.split()),
    }

    if raw_markdown:
        OUTPUTS_RAW.mkdir(parents=True, exist_ok=True)
        out_path = OUTPUTS_RAW / f"{path.stem}_raw.md"
        out_path.write_text(raw_markdown, encoding="utf-8")

    status = "failed" if not raw_markdown.strip() else "success"

    return {
        "source_type": "csv",
        "source_file": path.name,
        "ingestion_engine": "python-csv",
        "pdf_classification": classification,
        "raw_markdown": raw_markdown,
        "metadata": metadata,
        "ocr_required": False,
        "status": status,
        "errors": errors,
    }


def parse_tabular_to_markdown(file_path: str) -> dict:
    """Route supported table-like files to the right parser."""
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        return parse_csv_to_markdown(file_path)
    return parse_excel_to_markdown(file_path)
